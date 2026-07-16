import csv
import json
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from utils.validators import VALID_NUCLEI


class HistoricalImportError(ValueError):
    """Raised when a historical source file or row cannot be normalized."""


@dataclass(frozen=True)
class HistoricalProjectRecord:
    source_id: str
    area: str
    costs: dict[str, float]
    final_price: float
    project_date: date
    project_name: str = ""
    service: str = ""
    complexity: str = ""
    result: str = ""
    observations: str = ""

    def to_schema_dict(self) -> dict[str, Any]:
        return {
            "id": self.source_id,
            "area": self.area,
            "custos": self.costs,
            "precoFinalPraticado": self.final_price,
            "data": self.project_date.isoformat(),
            "nomeProjeto": self.project_name,
            "servico": self.service,
            "complexidade": self.complexity,
            "resultado": self.result,
            "observacoes": self.observations,
        }


@dataclass(frozen=True)
class ImportIssue:
    row_number: int
    reason: str


@dataclass
class HistoricalImportReport:
    records: list[HistoricalProjectRecord] = field(default_factory=list)
    issues: list[ImportIssue] = field(default_factory=list)
    total_rows: int = 0

    @property
    def summary(self) -> str:
        return (
            f"{len(self.records)} linhas importadas, "
            f"{len(self.issues)} ignoradas de {self.total_rows} processadas"
        )


def import_historical_file(
    source_path: str | Path,
    aliases_path: str | Path,
) -> HistoricalImportReport:
    source = Path(source_path)
    aliases = load_aliases(aliases_path)
    rows = read_tabular_rows(source)
    report = HistoricalImportReport(total_rows=len(rows))
    seen_source_ids: set[str] = set()

    for row_number, row in rows:
        try:
            record = normalize_historical_row(row, aliases)
            if record.source_id in seen_source_ids:
                raise HistoricalImportError(
                    f"duplicate project id in source file: {record.source_id}"
                )
            seen_source_ids.add(record.source_id)
            report.records.append(record)
        except HistoricalImportError as error:
            report.issues.append(ImportIssue(row_number=row_number, reason=str(error)))

    return report


def load_aliases(aliases_path: str | Path) -> dict[str, Any]:
    try:
        with Path(aliases_path).open(encoding="utf-8") as aliases_file:
            aliases = json.load(aliases_file)
    except (OSError, json.JSONDecodeError) as error:
        raise HistoricalImportError(f"could not load aliases file: {error}") from error

    if not isinstance(aliases, dict):
        raise HistoricalImportError("aliases file must contain a JSON object")
    return aliases


def read_tabular_rows(source_path: Path) -> list[tuple[int, dict[str, Any]]]:
    if not source_path.is_file():
        raise HistoricalImportError(f"historical source file not found: {source_path}")

    suffix = source_path.suffix.casefold()
    if suffix == ".csv":
        return read_csv_rows(source_path)
    if suffix == ".xlsx":
        return read_xlsx_rows(source_path)

    raise HistoricalImportError(
        f"unsupported historical file format {suffix or '<none>'}: expected .csv or .xlsx"
    )


def read_csv_rows(source_path: Path) -> list[tuple[int, dict[str, Any]]]:
    with source_path.open(encoding="utf-8-sig", newline="") as source_file:
        sample = source_file.read(4096)
        source_file.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel

        reader = csv.DictReader(source_file, dialect=dialect)
        if not reader.fieldnames:
            raise HistoricalImportError("historical CSV has no header row")

        rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, row in enumerate(reader, start=2):
            normalized_row = normalize_headers(row)
            if has_row_values(normalized_row):
                rows.append((row_number, normalized_row))
        return rows


def read_xlsx_rows(source_path: Path) -> list[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        values = worksheet.iter_rows(values_only=True)
        headers = next(values, None)
        if not headers:
            raise HistoricalImportError("historical XLSX has no header row")

        rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, values_row in enumerate(values, start=2):
            raw_row = dict(zip(headers, values_row, strict=False))
            normalized_row = normalize_headers(raw_row)
            if has_row_values(normalized_row):
                rows.append((row_number, normalized_row))
        return rows
    finally:
        workbook.close()


def normalize_headers(row: dict[Any, Any]) -> dict[str, Any]:
    return {
        normalize_key(str(header)): value
        for header, value in row.items()
        if header is not None
    }


def has_row_values(row: dict[str, Any]) -> bool:
    return any(value not in (None, "") for value in row.values())


def normalize_historical_row(
    row: dict[str, Any],
    aliases: dict[str, Any],
) -> HistoricalProjectRecord:
    source_id = require_text(row, aliases, "id")
    area = normalize_area(require_text(row, aliases, "area"))
    final_price = parse_nonnegative_number(
        require_value(row, aliases, "precoFinalPraticado"),
        "precoFinalPraticado",
    )
    project_date = parse_project_date(require_value(row, aliases, "data"))
    costs = parse_costs(row, aliases.get("custos", {}))

    return HistoricalProjectRecord(
        source_id=source_id,
        area=area,
        costs=costs,
        final_price=final_price,
        project_date=project_date,
        project_name=optional_text(row, aliases, "nomeProjeto"),
        service=optional_text(row, aliases, "servico"),
        complexity=optional_text(row, aliases, "complexidade"),
        result=optional_text(row, aliases, "resultado"),
        observations=optional_text(row, aliases, "observacoes"),
    )


def parse_costs(row: dict[str, Any], cost_aliases: Any) -> dict[str, float]:
    if not isinstance(cost_aliases, dict):
        raise HistoricalImportError("custos aliases must be a JSON object")

    costs: dict[str, float] = {}
    for cost_id, aliases in cost_aliases.items():
        value = find_value(row, aliases)
        if value in (None, ""):
            continue
        costs[cost_id] = parse_nonnegative_number(value, f"custos.{cost_id}")
    return costs


def require_text(row: dict[str, Any], aliases: dict[str, Any], field_name: str) -> str:
    value = require_value(row, aliases, field_name)
    text = str(value).strip()
    if not text:
        raise HistoricalImportError(f"missing required field: {field_name}")
    return text


def require_value(row: dict[str, Any], aliases: dict[str, Any], field_name: str) -> Any:
    value = find_value(row, aliases.get(field_name, []))
    if value in (None, ""):
        raise HistoricalImportError(f"missing required field: {field_name}")
    return value


def optional_text(row: dict[str, Any], aliases: dict[str, Any], field_name: str) -> str:
    value = find_value(row, aliases.get(field_name, []))
    return "" if value in (None, "") else str(value).strip()


def find_value(row: dict[str, Any], aliases: Any) -> Any:
    if not isinstance(aliases, list):
        return None

    for alias in aliases:
        value = row.get(normalize_key(str(alias)))
        if value not in (None, ""):
            return value
    return None


def normalize_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = "".join(character for character in normalized if not unicodedata.combining(character))
    return re.sub(r"[^a-z0-9]+", "_", ascii_value.casefold()).strip("_")


def normalize_area(value: str) -> str:
    areas_by_key = {normalize_key(area): area for area in VALID_NUCLEI}
    area = areas_by_key.get(normalize_key(value))
    if area is None:
        expected = ", ".join(sorted(VALID_NUCLEI))
        raise HistoricalImportError(f"invalid area {value!r}: expected one of {expected}")
    return area


def parse_nonnegative_number(value: Any, field_name: str) -> float:
    if isinstance(value, (int, float, Decimal)):
        decimal_value = Decimal(str(value))
    else:
        raw_value = re.sub(r"[^0-9,.-]", "", str(value).strip())
        if "," in raw_value and "." in raw_value:
            raw_value = (
                raw_value.replace(".", "").replace(",", ".")
                if raw_value.rfind(",") > raw_value.rfind(".")
                else raw_value.replace(",", "")
            )
        elif "," in raw_value:
            raw_value = raw_value.replace(".", "").replace(",", ".")

        try:
            decimal_value = Decimal(raw_value)
        except InvalidOperation as error:
            raise HistoricalImportError(
                f"invalid numeric field {field_name}: {value!r}"
            ) from error

    if decimal_value < 0:
        raise HistoricalImportError(f"negative value is not allowed for {field_name}")
    return float(decimal_value)


def parse_project_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    for date_format in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    raise HistoricalImportError(
        f"invalid date {value!r}: expected YYYY-MM-DD or DD/MM/YYYY"
    )
